"""Excel (.xlsx) parser using openpyxl in read-only mode."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from ..schemas import FilePreview, TablePreview
from .base import ParseError, Parser

_PREVIEW_ROWS = 100
_PREVIEW_COLS = 50


class ExcelParser(Parser):
    kind = "excel"

    def _open(self, data: bytes):
        try:
            return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            raise ParseError(f"Failed to read .xlsx: {e}") from e

    def meta(self, data: bytes, *, filename: str) -> dict[str, Any]:
        wb = self._open(data)
        try:
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheets.append({
                    "name": name,
                    "rows": ws.max_row or 0,
                    "cols": ws.max_column or 0,
                })
            return {"kind": self.kind, "sheets": sheets}
        finally:
            wb.close()

    def preview(self, data: bytes, *, filename: str) -> FilePreview:
        wb = self._open(data)
        try:
            first = wb.sheetnames[0] if wb.sheetnames else None
            if not first:
                return FilePreview(kind=self.kind, error="Empty workbook")
            ws = wb[first]
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0

            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                return FilePreview(
                    kind=self.kind,
                    table=TablePreview(
                        sheet_name=first, columns=[], rows=[],
                        total_rows=0, total_cols=0,
                    ),
                )
            columns = [str(c) if c is not None else "" for c in header_row[:_PREVIEW_COLS]]

            rows: list[list[Any]] = []
            for i, row in enumerate(it):
                if i >= _PREVIEW_ROWS:
                    break
                rows.append([_jsonable(v) for v in row[:_PREVIEW_COLS]])

            return FilePreview(
                kind=self.kind,
                truncated=(total_rows - 1 > _PREVIEW_ROWS) or (total_cols > _PREVIEW_COLS),
                table=TablePreview(
                    sheet_name=first,
                    columns=columns,
                    rows=rows,
                    total_rows=max(total_rows - 1, 0),  # excluding header
                    total_cols=total_cols,
                ),
            )
        finally:
            wb.close()


def _jsonable(v: Any) -> Any:
    """Make openpyxl values JSON-serialisable. Datetimes → isoformat."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    iso = getattr(v, "isoformat", None)
    if callable(iso):
        try:
            return iso()
        except Exception:
            pass
    return str(v)
